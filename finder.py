import argparse
import json
import re
import time
from datetime import datetime, timezone
from urllib.parse import urlencode
from bs4 import BeautifulSoup
import boto3
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
import io
from collections import defaultdict

# ---- Config -----------------------------------------------------------------
BUCKET  = "p000268ds-comp-offers"
TIER_JSON    = "json"
TIER_EXCEL   = "excel"
DATASET = "offers"
SOURCE  = "finder"
URL     = "https://www.finder.com.au/health-insurance"
SHEET = "comp_offer.xlsx" # competition offer table filename (for excel output)

BRANDS = {
    "medibank": "Medibank",
    "ahm":      "ahm",
    "bupa":     "Bupa",
    "nib":      "nib",
    "hcf":      "HCF",
    "hbf":      "HBF",
}

# Shared quiz params — encode the "average Australian" profile
QUIZ_PARAMS = {
    "lifestage":   "Any",
    "state":       "VIC",
    "INCOME_TIER": "base_tier",
    "DOB_MAIN":    "1990-01-01",
    "quizId":      "25f6d5b4-be64-46da-b4ce-1049bbd4a546",
}

# Per-cover-type params — only what differs from QUIZ_PARAMS.
# covercategory is omitted for extras only
COVER_CATEGORY = ["basic", "bronze", "silver", "gold"]

COVER_TYPES = {
    "hospital + extras": {"covertype": "Combined", "covercategory": COVER_CATEGORY},
    "hospital only":     {"covertype": "Hospital", "covercategory": COVER_CATEGORY},
    "extras only":       {"covertype": "Extras"}
}

BASE_RESULTS_URL = "https://www.finder.com.au/health-insurance/health-insurance-results"
FINDER_REWARDS_URL = "https://www.finder.com.au/finder-rewards"

# ---- Compiled regex patterns ------------------------------------------------
WEEKS_FREE_PAT  = re.compile(r'\d+\s*(?:\+\d+\s*)?weeks?\s*free', re.I)
WAITING_PAT     = re.compile(r'\d+\s*(?:and\s*\d+\s*)?month.*?wait|waived waiting|no waiting period|waits waived', re.I)
GIFT_CARD_PAT   = re.compile(r'\$\d+\s*[\w\s]*?\b(?:gift\s*card|e-gift)', re.I)
OTHER_KEYWORDS  = re.compile(r'loyalty|reward|discount|bonus|cashback|cash\s*back|voucher|prize|store|e-gift', re.I)
PRICE_PREFIX_PAT = re.compile(r'^\$?\d+(?:\.\d+)?\s*(?:per\s+\w+\s*)?', re.I)
SENTENCE_SPLIT  = re.compile(r'(?<=[.!?])\s+')

BOILERPLATE_PATTERNS = [
        r'Go to Site',
        r'View details',
        r'Compare product selection',
        r'Compare loading',
        r'loading',
        r'\[View details\]',
        r'\|\s*loading\s*\|']

REWARDS_COVER = {
    "hospital & extras":  "hospital + extras",
    "hospital + extras":  "hospital + extras", # 2 aliases for same cover type since the site uses both terms inconsistently
    "hospital only":      "hospital only",
    "extras only":        "extras only",
}

EXCEL_COL_MAP = {
    "weeks_free": "Offer : Weeks Free",
    "waiting_waive": "Offer : Waiting period waive",
    "other": "Offer : Other",
    "end_date": "Offer : End date",
}

#---- URL Builder -------------------------------------------------------------
def build_url(cover_params: dict) -> str:
    """Build results URL by merging shared QUIZ_PARAMS with cover-specific params."""
    params = {**QUIZ_PARAMS, **cover_params}
    return f"{BASE_RESULTS_URL}?{urlencode(params)}#quiz-results-table"

# ---- Helpers ----------------------------------------------------------------
def make_browser_context(playwright):
    browser = playwright.chromium.launch(
        headless=True,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
    )
    ctx = browser.new_context(
        user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        viewport={"width": 1280, "height": 900},
        locale="en-AU",
        timezone_id="Australia/Melbourne",
        extra_http_headers={"Accept-Language": "en-AU,en;q=0.9"},
    )
    ctx.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return browser, ctx


# ---- Core Functions for Scraoing ---------------------------------------------
def load_results_page(page, cover_params: dict):
    url = build_url(cover_params)
    page.goto(url, timeout=30000, wait_until="domcontentloaded")
    time.sleep(3)

    try:
        page.click("button:has-text('Accept')", timeout=4000)
        time.sleep(1)
    except PWTimeout:
        pass

    try:
        page.wait_for_selector("table tr", timeout=15000)
    except PWTimeout:
        pass


def reset_provider_filters(page):
    try:
        for text in ["All providers", "All Providers", "All health insurers", "All"]:
            all_label = page.query_selector(f"label:has-text('{text}')")
            if all_label:
                all_label.scroll_into_view_if_needed()
                all_label.click()
                time.sleep(2)
                return True
    except Exception:
        pass
    return False


def filter_by_provider(page, brand_label: str) -> bool:
    try:
        reset_provider_filters(page)

        search_input = page.wait_for_selector("input.searchInput[placeholder='Search for a provider']", timeout=10000)
        search_input.click()
        search_input.fill("")
        search_input.type(brand_label, delay=80)
        time.sleep(3)

        label = (page.query_selector(f"label:has-text('{brand_label}')") or
                 page.query_selector(f"label:has-text('{brand_label.lower()}')"))

        if not label:
            return False

        label.scroll_into_view_if_needed()
        label.click()
        time.sleep(4)
        return True
    except Exception:
        return False


def clear_provider_filter(page, url: str):
    page.goto(url, timeout=30000, wait_until="domcontentloaded")
    time.sleep(3)


def parse_offer_text(raw_text: str) -> dict:

    for pattern in BOILERPLATE_PATTERNS:
        # Remove boilerplate phrases that could interfere with offer parsing
        raw_text = re.sub(pattern, '', raw_text, flags=re.I)
    
    # Clean up extra spaces and pipes
    cleaned = re.sub(r'\s*\|\s*', ' | ', raw_text)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    lower = cleaned.lower()
 
    # Weeks free
    weeks = ""
    m = re.search(r'(?:up to\s+)?(\d+(?:\+\d+)?)\s*weeks?\s*free', lower)
    if m:
        weeks = m.group(1)
 
    # Waiting period waive
    waiting = ""
    m_wait = re.search(r'(\d+)\s*and\s*(\d+)\s*month', lower)
    if m_wait:
        waiting = f"{m_wait.group(1)} and {m_wait.group(2)} month waits waived"
    else:
        m_wait2 = re.search(r'(\d+)\s*month.*?(?:wait|extra)', lower)
        if m_wait2:
            waiting = f"{m_wait2.group(1)} month waits waived"
    if not waiting and any(x in lower for x in ["waived waiting", "no waiting period", "waits waived"]):
        waiting = "waiting period waived"
 
    # Offer end date
    end_date = ""
    date_match = re.search(r'(?:ends?|until|by)\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})', lower)
    if date_match:
        end_date = date_match.group(1)
    else:
        date_match2 = re.search(r'(\d{1,2}\s+[A-Za-z]{3,}\s+\d{4})', lower)
        if date_match2 and any(y in lower for y in ["offer ends", "ends", "join by"]):
            end_date = date_match2.group(1)
 
    # Gift card, extracted independently so it isn't missed when it shares a sentence with weeks free
    gift_card = ""
    gift_match = re.search(r'\$(\d+)\s*(?:[\w\s]*?)\b(?:gift\s*card|e-gift)', lower)
    if gift_match:
        gift_card = f"${gift_match.group(1)} gift card"
 
    # Other, full sentences containing offer keywords, excluding weeks-free, waiting-period,
    # and gift-card sentences (gift card is already captured above as a structured value).
    sentences = SENTENCE_SPLIT.split(raw_text.strip())
    other_sentences = []
    for sent in sentences:
        sent = sent.strip()
        if not sent:
            continue
        if WEEKS_FREE_PAT.search(sent):
            continue
        if WAITING_PAT.search(sent):
            continue
        if GIFT_CARD_PAT.search(sent):
            continue  # already captured as gift_card
        if OTHER_KEYWORDS.search(sent):
            cleaned = PRICE_PREFIX_PAT.sub("", sent).strip()
            if cleaned:
                other_sentences.append(cleaned)
 
    other_parts = []
    if gift_card:
        other_parts.append(gift_card)
    if other_sentences:
        other_parts.append(" ".join(other_sentences))
 
    return {
        "weeks_free": weeks,
        "waiting_waive": waiting,
        "other": " | ".join(other_parts),
        "end_date": end_date
    }


def extract_offers(page, cover_type_name: str, cover_params: dict) -> list:
    results = []
    base_url = build_url(cover_params)

    for brand_key, search_name in BRANDS.items():
        applied = filter_by_provider(page, search_name)
        if not applied:
            results.append({
                "brand": brand_key,
                "cover_type": cover_type_name,
                "cover_category": cover_params.get("covercategory", ""),
                "weeks_free": "",
                "waiting_waive": "",
                "other": "",
                "end_date": ""
            })
            clear_provider_filter(page, base_url)
            continue

        time.sleep(2)
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")
        rows = soup.select("table tr")

        offer_data = {
            "brand": brand_key,
            "cover_type": cover_type_name,
            "cover_category": cover_params.get("covercategory", ""),
            "weeks_free": "",
            "waiting_waive": "",
            "other": "",
            "end_date": ""
        }

        if rows and len(rows) > 1:
            first_row_text = rows[1].get_text(" ", strip=True)
            offer_data.update(parse_offer_text(first_row_text))

        results.append(offer_data)
        clear_provider_filter(page, base_url)
        time.sleep(1.5)

    return results

# ---- Finder Rewards Scraping and Annotation --------------------------------
def scrape_finder_rewards(page) -> list:
    """
    Load the Finder Rewards page and extract health-insurance-relevant rewards.
    """
    print("Scraping Finder Rewards page...")
    page.goto(FINDER_REWARDS_URL, timeout=30000, wait_until="domcontentloaded")
    time.sleep(3)
 
    try:
        page.click("button:has-text('Accept')", timeout=4000)
        time.sleep(1)
    except PWTimeout:
        pass
 
    html = page.content()
    soup = BeautifulSoup(html, "html.parser")
 
    rewards = []
    end_date_pat = re.compile(r'ends?\s+(\d{1,2}\s+[a-z]{3,}\s+\d{4})', re.I)
    get_with_pat = re.compile(r'Get\s+(?:up to\s+)?\$(\d+)\s+with\s+(.+)', re.I)
 
    for el in soup.find_all(string=get_with_pat):
        m = get_with_pat.search(el)
        if not m:
            continue
 
        amount_val = f"${m.group(1)}"
        brand_raw  = m.group(2).strip()
 
        brand_key = None
        for internal_key in BRANDS:
            if internal_key.lower() in brand_raw.lower():
                brand_key = internal_key
                break
        if not brand_key:
            continue
 
        # Walk up the DOM to find a container holding both category and end date
        container = el.parent
        for _ in range(6):
            container_text = container.get_text(" ", strip=True).lower()
            if "health insurance" in container_text and end_date_pat.search(container_text):
                break
            if container.parent:
                container = container.parent
            else:
                break
 
        container_lower = container.get_text(" ", strip=True).lower()
 
        if "health insurance" not in container_lower:
            continue
 
        ed_m = end_date_pat.search(container_lower)
        end_date = ed_m.group(1).strip() if ed_m else ""
 
        # Determine cover_type from category text; None = applies to all types
        cover_type = None
        for alias, ct in REWARDS_COVER.items():
            if alias.lower() in container_lower:
                cover_type = ct
                break
 
        rewards.append({
            "brand_key":  brand_key,
            "cover_type": cover_type,
            "amount":     amount_val,
            "end_date":   end_date,
        })
        print(f"  Found Finder Reward: {brand_key} | {cover_type} | {amount_val} | ends {end_date}")
 
    return rewards
 
 
def annotate_with_finder_rewards(offers: list, rewards: list) -> list:
    """
    For each offer, keyword-match on:
      1. brand  — offer["brand"] matches reward["brand_key"]
      2. amount — offer["other"] contains "$X REWARD" matching reward["amount"]
    """
    reward_pat_cache = {}
 
    for offer in offers:
        other_col = offer.get("other", "")
        end_date_col = offer.get("end_date", "") 
        if not other_col:
            continue
 
        for reward in rewards:
            if offer["brand"] != reward["brand_key"]:
                continue
 
            amount   = reward["amount"]
            end_date = reward["end_date"]
            tag      = f"Finder Rewards"
 
            if amount not in reward_pat_cache:
                reward_pat_cache[amount] = re.compile(
                    re.escape(amount) + r'\s*REWARD', re.I
                )
            pat = reward_pat_cache[amount]
 
            if pat.search(other_col) and tag not in other_col and tag not in end_date_col:
                other_col = pat.sub(lambda mo: f"{mo.group(0)} ({tag})", other_col)
                offer["other"] = other_col  # update for subsequent reward iterations
                offer["end_date"] = end_date_col + f" | {tag}: {end_date}" if end_date else end_date_col
 
    return offers
 
 
def run() -> list:
    all_offers = []
    with sync_playwright() as p:
        browser, ctx = make_browser_context(p)
        page = ctx.new_page()
 
        try:
            for cover_key, cover_params in COVER_TYPES.items():
                for covercategory in cover_params.get("covercategory", [""]):
                    print(f"Scraping {cover_key}...")
                    if covercategory:
                        cover_params["covercategory"] = covercategory
                        print(f"cover_params: {cover_params}")
                    load_results_page(page, cover_params)
                    offers = extract_offers(page, cover_key, cover_params)
                    all_offers.extend(offers)
 
            # Scrape Finder Rewards and annotate matching offers
            finder_rewards = scrape_finder_rewards(page)
            all_offers = annotate_with_finder_rewards(all_offers, finder_rewards)
 
        except Exception as e:
            print(f"[error] {e}")
            raise
        finally:
            browser.close()
 
    return all_offers

# ---- Output -----------------------------------------------------------------
def build_payload(content: list[dict]) -> dict:
    return {
        "source": SOURCE,
        "tier": TIER_JSON,
        "dataset": DATASET,
        "scraped_at": datetime.now(timezone.utc).isoformat(),
        "url": URL,
        "content": content
    }
 
 
def save_local(payload: dict, directory: str = ".") -> None:
    run_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    path = f"{directory}/{payload['source']}_{run_date}.json"
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
    print(f"Saved locally: {path}")
 
 
def upload_to_s3(payload: dict) -> None:
    s3 = boto3.client("s3", region_name="us-east-1")
    key = f"raw/{payload['tier']}/{payload['source']}_{payload['dataset']}_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.json"
    s3.put_object(
        Bucket=BUCKET,
        Key=key,
        Body=json.dumps(payload, ensure_ascii=False, indent=2),
        ContentType="application/json",
    )
    print(f"Uploaded: s3://{BUCKET}/{key}")

# ---- Excel helpers ----------------------------------------------------------
 
# Maps offer field → exact column header name from the excel sheet
EXCEL_COL_MAP = {
    "weeks_free": "Offer : Weeks Free",
    "waiting_waive": "Offer : Waiting period waive",
    "other": "Offer : Other",
    "end_date": "Offer : End date",
}
 
 
def _header_index(ws) -> dict[str, int]:
    """Return {column_name: 0-based col index} by scanning all header rows."""
    idx = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() in EXCEL_COL_MAP.values():
                idx[cell.value.strip()] = cell.column - 1
        if len(idx) == len(EXCEL_COL_MAP):
            break
    return idx

def cat_sort_key(cat: str) -> int:
    try:
        return COVER_CATEGORY.index(cat)
    except ValueError:
        return 99

def format_cats(cats: list[str]) -> str:
    return "/".join(sorted(cats, key=cat_sort_key))


def collapse_categories(category_records: list[dict]) -> dict:
    """Collapse multiple category offers into compact strings.

    For each field, if all categories share the same value, emit it once.
    If values differ, prefix each distinct value with the categories that share it.

    Example output for weeks_free when identical:
        "basic/bronze/silver/gold: 12"
    Example output when split:
        "basic/bronze/gold: 10 | silver: 12"
    """
    if not category_records:
        return {}

    if len(category_records) == 1 or all(r.get("cover_category") == "all" for r in category_records):
        return category_records[0]

    result = dict(category_records[0])  # start from first record as base

    for field in ["weeks_free", "waiting_waive", "other", "end_date"]:
        # Map each category → its value for this field (empty string if missing)
        cat_to_val: dict[str, str] = {
            r.get("cover_category", "n/a"): r.get(field, "")
            for r in category_records
            if r.get("cover_category") != "all"
        }

        # If no record has a value, leave blank
        if not any(cat_to_val.values()):
            result[field] = ""
            continue

        # Group categories by their value for this field
        val_to_cats: dict[str, list[str]] = defaultdict(list)
        for cat, val in cat_to_val.items():
            val_to_cats[val].append(cat)

        unique_vals = set(cat_to_val.values())

        # All identical (including all-empty handled above, so here all non-empty & same)
        if len(unique_vals) == 1:
            result[field] = unique_vals.pop()
            continue

        # Mixed values — emit "cats: value" for each non-empty group, sorted by lowest cat tier
        parts = []
        for val, cats in sorted(val_to_cats.items(),
                                key=lambda kv: cat_sort_key(min(kv[1], key=cat_sort_key))):
            if val:  # skip the empty-value group
                parts.append(f"{format_cats(cats)}: {val}")

        result[field] = " | ".join(parts) if parts else ""

    return result


def fill_excel(wb: openpyxl.Workbook, offers: list[dict]) -> None:
    """Write Aggregator offer data with collapsed categories."""
    ws = wb["table"]
    col_idx = _header_index(ws)
    
    # Group offers by (brand, cover_type)
    grouped = defaultdict(list)
    for offer in offers:
        if offer.get("brand") not in BRANDS:
            continue
        key = (offer["brand"].lower(), offer["cover_type"].lower())
        grouped[key].append(offer)

    for row in ws.iter_rows(min_row=2):
        if row[2].value != "Aggregator":   # Channel column
            continue
            
        brand_val = str(row[0].value or "").strip().lower()
        cover_val = str(row[1].value or "").strip().lower()
        
        if not brand_val or not cover_val:
            continue
            
        key = (brand_val, cover_val)
        category_offers = grouped.get(key)
        
        if not category_offers:
            continue

        # Collapse categories
        collapsed = collapse_categories(category_offers)
        
        # Fill the row
        for field, col_name in EXCEL_COL_MAP.items():
            if col_name in col_idx:
                value = collapsed.get(field)
                row[col_idx[col_name]].value = value if value else None
 
 
def update_on_s3(offers: list[dict]) -> None:
    """Download comp_offer.xlsx from S3, fill Aggregator rows, re-upload (overwrite)."""
    s3 = boto3.client("s3", region_name="us-east-1")
    key = f"raw/{TIER_EXCEL}/{SHEET}"
    obj = s3.get_object(Bucket=BUCKET, Key=key)
    wb = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()))
    fill_excel(wb, offers)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3.put_object(
        Bucket=BUCKET,
        Key=key,
        Body=buf.read(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    print(f"Updated: s3://{BUCKET}/{key}")
 
 
def update_excel_local(offers: list[dict], directory: str = ".") -> None:
    """Fill Aggregator rows in a local copy of comp_offer.xlsx and save in place."""
    src = f"{directory}/{SHEET}"
    wb = openpyxl.load_workbook(src)
    fill_excel(wb, offers)
    wb.save(src)
    print(f"Updated locally: {src}")
 
 
# ---- Main -------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Finder Health Insurance Scraper")
    parser.add_argument("--local", metavar="DIR", nargs="?", const=".", help="Save locally")
    args = parser.parse_args()

    content_list = run()
    payload = build_payload(content_list)

    if args.local:
        save_local(payload, args.local)
        update_excel_local(content_list, args.local)
    else:
        upload_to_s3(payload)
        update_on_s3(content_list)
